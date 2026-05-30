from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QPushButton, QLineEdit, QComboBox, QFileDialog, QMessageBox, QDialog,
    QFormLayout, QDialogButtonBox, QTextEdit, QMenu
)
from PyQt6.QtCore import Qt
from database import Session
from models import Customer
import pandas as pd


class CustomerDialog(QDialog):
    """نافذة إضافة/تعديل عميل"""
    
    def __init__(self, parent=None, customer=None):
        super().__init__(parent)
        self.customer = customer
        self.setWindowTitle("Add Customer" if not customer else "Edit Customer")
        self.setMinimumWidth(400)
        
        layout = QFormLayout()
        
        self.name_input = QLineEdit()
        self.email_input = QLineEdit()
        self.phone_input = QLineEdit()
        self.company_input = QLineEdit()
        self.status_combo = QComboBox()
        self.status_combo.addItems(["lead", "prospect", "active", "churned"])
        self.notes_input = QTextEdit()
        self.notes_input.setMaximumHeight(80)
        
        layout.addRow("Name*:", self.name_input)
        layout.addRow("Email:", self.email_input)
        layout.addRow("Phone:", self.phone_input)
        layout.addRow("Company:", self.company_input)
        layout.addRow("Status:", self.status_combo)
        layout.addRow("Notes:", self.notes_input)
        
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)
        
        self.setLayout(layout)
        
        if customer:
            self.name_input.setText(customer.name)
            self.email_input.setText(customer.email or "")
            self.phone_input.setText(customer.phone or "")
            self.company_input.setText(customer.company or "")
            self.status_combo.setCurrentText(customer.status)
            self.notes_input.setText(customer.notes or "")
    
    def get_data(self):
        return {
            'name': self.name_input.text().strip(),
            'email': self.email_input.text().strip(),
            'phone': self.phone_input.text().strip(),
            'company': self.company_input.text().strip(),
            'status': self.status_combo.currentText(),
            'notes': self.notes_input.toPlainText().strip()
        }


class CustomersTab(QWidget):
    def __init__(self):
        super().__init__()
        self.session = Session()
        self.init_ui()
        self.load_data()

    def init_ui(self):
        controls_layout = QHBoxLayout()
        
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search by name, email...")
        self.search_input.textChanged.connect(self.filter_data)
        controls_layout.addWidget(self.search_input)
        
        self.status_filter = QComboBox()
        self.status_filter.addItems(["All", "Lead", "Prospect", "Active", "Churned"])
        self.status_filter.currentIndexChanged.connect(self.filter_data)
        controls_layout.addWidget(self.status_filter)
        
        self.add_btn = QPushButton("Add Customer")
        self.add_btn.clicked.connect(self.add_customer)
        controls_layout.addWidget(self.add_btn)
        
        self.edit_btn = QPushButton("Edit")
        self.edit_btn.clicked.connect(self.edit_customer)
        controls_layout.addWidget(self.edit_btn)
        
        self.delete_btn = QPushButton("Delete")
        self.delete_btn.clicked.connect(self.delete_customer)
        controls_layout.addWidget(self.delete_btn)
        
        self.export_btn = QPushButton("Export")
        self.export_btn.clicked.connect(self.show_export_menu)
        controls_layout.addWidget(self.export_btn)
        
        self.import_btn = QPushButton("Import")
        self.import_btn.clicked.connect(self.import_data)
        controls_layout.addWidget(self.import_btn)
        
        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels(["ID", "Name", "Email", "Phone", "Company", "Status", "Deals"])
        self.table.setSortingEnabled(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.doubleClicked.connect(self.edit_customer)
        self.table.setColumnHidden(0, True)
        
        main_layout = QVBoxLayout()
        main_layout.addLayout(controls_layout)
        main_layout.addWidget(self.table)
        self.setLayout(main_layout)

    def load_data(self):
        customers = self.session.query(Customer).all()
        
        self.table.setRowCount(len(customers))
        for row, customer in enumerate(customers):
            self.table.setItem(row, 0, QTableWidgetItem(str(customer.id)))
            self.table.setItem(row, 1, QTableWidgetItem(customer.name))
            self.table.setItem(row, 2, QTableWidgetItem(customer.email or ""))
            self.table.setItem(row, 3, QTableWidgetItem(customer.phone or ""))
            self.table.setItem(row, 4, QTableWidgetItem(customer.company or ""))
            self.table.setItem(row, 5, QTableWidgetItem(customer.status))
            self.table.setItem(row, 6, QTableWidgetItem(str(len(customer.deals))))

    def filter_data(self):
        search_text = self.search_input.text().lower()
        status_filter = self.status_filter.currentText()
        
        for row in range(self.table.rowCount()):
            name = self.table.item(row, 1).text().lower()
            email = self.table.item(row, 2).text().lower()
            status = self.table.item(row, 5).text()
            
            show = True
            if search_text and search_text not in name and search_text not in email:
                show = False
            
            if status_filter != "All" and status != status_filter.lower():
                show = False
            
            self.table.setRowHidden(row, not show)

    def add_customer(self):
        dialog = CustomerDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            data = dialog.get_data()
            
            if not data['name']:
                QMessageBox.warning(self, "Error", "Name is required!")
                return
            
            customer = Customer(**data)
            self.session.add(customer)
            self.session.commit()
            
            QMessageBox.information(self, "Success", "Customer added successfully!")
            self.load_data()

    def edit_customer(self):
        selected = self.table.currentRow()
        if selected < 0:
            QMessageBox.warning(self, "Edit", "Please select a customer to edit")
            return
        
        customer_id = int(self.table.item(selected, 0).text())
        customer = self.session.get(Customer, customer_id)
        
        dialog = CustomerDialog(self, customer)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            data = dialog.get_data()
            
            for key, value in data.items():
                setattr(customer, key, value)
            
            self.session.commit()
            QMessageBox.information(self, "Success", "Customer updated successfully!")
            self.load_data()

    def delete_customer(self):
        selected = self.table.currentRow()
        if selected < 0:
            QMessageBox.warning(self, "Delete", "Please select a customer to delete")
            return
        
        reply = QMessageBox.question(
            self,
            "Delete Customer",
            "Are you sure you want to delete this customer?\nThis will also delete all associated deals!",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            customer_id = int(self.table.item(selected, 0).text())
            customer = self.session.get(Customer, customer_id)
            self.session.delete(customer)
            self.session.commit()
            
            QMessageBox.information(self, "Success", "Customer deleted successfully!")
            self.load_data()

    def show_export_menu(self):
        menu = QMenu(self)
        menu.addAction("Excel (XLSX)", lambda: self.export_data('xlsx'))
        # ❌ تمت إزالة XLS
        menu.addAction("Excel Template (XLTX)", lambda: self.export_data('xltx'))
        menu.addAction("CSV", lambda: self.export_data('csv'))
        menu.addAction("PDF Report", lambda: self.export_data('pdf'))
        menu.exec(self.export_btn.mapToGlobal(self.export_btn.rect().bottomLeft()))

    def export_data(self, format_type):
        customers = self.session.query(Customer).all()
        
        if not customers:
            QMessageBox.warning(self, "Export", "No customers to export!")
            return
        
        data = []
        for c in customers:
            data.append({
                'Name': c.name,
                'Email': c.email or '',
                'Phone': c.phone or '',
                'Company': c.company or '',
                'Status': c.status,
                'Notes': c.notes or '',
                'Deals Count': len(c.deals)
            })
        
        df = pd.DataFrame(data)
        
        # ❌ تمت إزالة XLS من القائمة
        file_types = {
            'xlsx': "Excel Files (*.xlsx)",
            'xltx': "Excel Template (*.xltx)",
            'csv': "CSV Files (*.csv)",
            'pdf': "PDF Files (*.pdf)"
        }
        
        path, _ = QFileDialog.getSaveFileName(
            self, "Save File", f"customers.{format_type}", file_types[format_type]
        )
        
        if not path:
            return
        
        try:
            # CSV
            if format_type == 'csv':
                df.to_csv(path, index=False, encoding='utf-8-sig')
            
            # XLSX
            elif format_type == 'xlsx':
                df.to_excel(path, index=False, engine='openpyxl')
            
            # XLTX (Excel Template)
            elif format_type == 'xltx':
                from openpyxl import Workbook
                from openpyxl.utils.dataframe import dataframe_to_rows
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Customers"
                
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)
                
                wb.template = True
                wb.save(path)
            
            # PDF
            elif format_type == 'pdf':
                from fpdf import FPDF
                
                pdf = FPDF()
                pdf.add_page()
                pdf.set_font("Arial", "B", 16)
                pdf.cell(0, 10, "Customers Report", ln=True, align="C")
                pdf.ln(5)
                
                pdf.set_font("Arial", "B", 9)
                col_width = 190 / len(df.columns)
                for col in df.columns:
                    pdf.cell(col_width, 8, str(col), 1, 0, "C")
                pdf.ln()
                
                pdf.set_font("Arial", "", 8)
                for _, row in df.iterrows():
                    for col in df.columns:
                        val = str(row[col])[:25] if row[col] is not None else ""
                        pdf.cell(col_width, 6, val, 1, 0, "L")
                    pdf.ln()
                
                pdf.output(path)
            
            QMessageBox.information(self, "Export Successful", f"File saved to:\n{path}")
        
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export:\n{str(e)}")

    def import_data(self):
        # ✅ تم إزالة XLS من قائمة الاستيراد
        path, _ = QFileDialog.getOpenFileName(
            self, "Import Customers", "",
            "All Supported (*.xlsx *.xltx *.csv);;Excel (*.xlsx *.xltx);;CSV (*.csv)"
        )
        
        if not path:
            return
        
        try:
            # قراءة الملف
            if path.lower().endswith('.csv'):
                df = pd.read_csv(path, encoding='utf-8-sig')
            else:
                df = pd.read_excel(path, engine='openpyxl')
            
            imported = 0
            errors = []
            
            for idx, row in df.iterrows():
                try:
                    if pd.isna(row.get("Name")) or str(row.get("Name", "")).strip() == "":
                        errors.append(f"Row {idx+2}: Name is required")
                        continue
                    
                    customer = Customer(
                        name=str(row.get("Name", "")).strip(),
                        email=str(row.get("Email", "")).strip() if pd.notna(row.get("Email")) else None,
                        phone=str(row.get("Phone", "")).strip() if pd.notna(row.get("Phone")) else None,
                        company=str(row.get("Company", "")).strip() if pd.notna(row.get("Company")) else None,
                        status=str(row.get("Status", "lead")).strip().lower(),
                        notes=str(row.get("Notes", "")).strip() if pd.notna(row.get("Notes")) else None
                    )
                    
                    if customer.status not in ['lead', 'prospect', 'active', 'churned']:
                        customer.status = 'lead'
                    
                    self.session.add(customer)
                    imported += 1
                    
                except Exception as e:
                    errors.append(f"Row {idx+2}: {str(e)}")
            
            self.session.commit()
            
            msg = f"Successfully imported {imported} customers!"
            if errors:
                msg += f"\n\nErrors ({len(errors)}):\n" + "\n".join(errors[:10])
                if len(errors) > 10:
                    msg += f"\n... and {len(errors)-10} more errors"
            
            QMessageBox.information(self, "Import Complete", msg)
            self.load_data()
            
        except Exception as e:
            self.session.rollback()
            QMessageBox.critical(self, "Import Error", f"Failed to import file:\n{str(e)}")