import sys
import sqlite3
import os
import re
import csv
from datetime import datetime, timedelta

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QTabWidget, QWidget, QVBoxLayout, QHBoxLayout,
    QTableView, QPushButton, QLineEdit, QLabel, QFormLayout,
    QHeaderView, QMessageBox, QGroupBox, QComboBox, QDateEdit,
    QDialog, QListWidget, QTextEdit, QSplitter, QInputDialog,
    QFileDialog, QListWidgetItem, QCalendarWidget, QFrame, QScrollArea, QGridLayout,
    QAbstractItemView, QSizePolicy
)
from PyQt6.QtCore import Qt, QAbstractTableModel, QDate, QEvent, QTimer, QLocale
from PyQt6.QtGui import QFont, QSyntaxHighlighter, QTextCharFormat, QColor, QTextCursor

# --------------------------- DATABASE ---------------------------
class Database:
    def __init__(self, db_name="crm_data.db"):
        self.conn = sqlite3.connect(db_name)
        self.create_tables()
        self.migrate_if_needed()

    def create_tables(self):
        self.conn.execute("""
            CREATE TABLE IF NOT EXISTS activity_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                action TEXT NOT NULL,
                entity TEXT NOT NULL,
                entity_id INTEGER,
                details TEXT,
                timestamp TEXT NOT NULL DEFAULT (datetime('now','localtime'))
            )
        """)
        self.conn.execute("""
            CREATE TABLE IF NOT EXISTS companies (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                address TEXT,
                phone TEXT,
                website TEXT,
                notes TEXT
            )
        """)
        self.conn.execute("""
            CREATE TABLE IF NOT EXISTS contacts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                phone TEXT,
                email TEXT,
                company_id INTEGER,
                notes TEXT,
                FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE SET NULL
            )
        """)
        self.conn.execute("""
            CREATE TABLE IF NOT EXISTS attachments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                contact_id INTEGER NOT NULL,
                filename TEXT NOT NULL,
                data BLOB,
                FOREIGN KEY(contact_id) REFERENCES contacts(id) ON DELETE CASCADE
            )
        """)
        self.conn.execute("""
            CREATE TABLE IF NOT EXISTS communication_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                contact_id INTEGER NOT NULL,
                type TEXT NOT NULL,
                date TEXT NOT NULL,
                description TEXT,
                FOREIGN KEY(contact_id) REFERENCES contacts(id) ON DELETE CASCADE
            )
        """)
        self.conn.execute("""
            CREATE TABLE IF NOT EXISTS tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                description TEXT,
                due_date TEXT,
                status TEXT DEFAULT 'Pending',
                contact_id INTEGER,
                FOREIGN KEY(contact_id) REFERENCES contacts(id) ON DELETE SET NULL
            )
        """)
        self.conn.execute("""
            CREATE TABLE IF NOT EXISTS deals (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                value REAL DEFAULT 0.0,
                stage TEXT DEFAULT 'Lead',
                expected_close_date TEXT,
                contact_id INTEGER,
                FOREIGN KEY(contact_id) REFERENCES contacts(id) ON DELETE SET NULL
            )
        """)
        self.conn.execute("""
            CREATE TABLE IF NOT EXISTS scripts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                content TEXT
            )
        """)
        self.conn.commit()

    def migrate_if_needed(self):
        try:
            self.conn.execute("SELECT company_id FROM contacts LIMIT 1")
        except sqlite3.OperationalError:
            self.conn.execute("ALTER TABLE contacts ADD COLUMN company_id INTEGER REFERENCES companies(id)")
            self.conn.commit()

    # ========== Activity Log ==========
    def log_activity(self, action, entity, entity_id=None, details=""):
        self.conn.execute(
            "INSERT INTO activity_log (action, entity, entity_id, details) VALUES (?, ?, ?, ?)",
            (action, entity, entity_id, details)
        )
        self.conn.commit()

    def fetch_recent_activities(self, limit=100):
        cur = self.conn.execute("""
            SELECT timestamp, action, entity, details
            FROM activity_log
            ORDER BY id DESC LIMIT ?
        """, (limit,))
        return cur.fetchall()

    # ========== Contacts ==========
    def fetch_all_contacts(self):
        cur = self.conn.execute("""
            SELECT c.id, c.name, c.phone, c.email, co.name, c.notes
            FROM contacts c LEFT JOIN companies co ON c.company_id = co.id
            ORDER BY c.name
        """)
        return cur.fetchall()

    def search_contacts(self, keyword):
        cur = self.conn.execute("""
            SELECT c.id, c.name, c.phone, c.email, co.name, c.notes
            FROM contacts c LEFT JOIN companies co ON c.company_id = co.id
            WHERE c.name LIKE ? OR c.phone LIKE ? OR c.email LIKE ? OR co.name LIKE ?
            ORDER BY c.name
        """, (f'%{keyword}%', f'%{keyword}%', f'%{keyword}%', f'%{keyword}%'))
        return cur.fetchall()

    def add_contact(self, name, phone, email, company_id, notes):
        self.conn.execute(
            "INSERT INTO contacts (name, phone, email, company_id, notes) VALUES (?, ?, ?, ?, ?)",
            (name, phone, email, company_id, notes)
        )
        self.conn.commit()
        self.log_activity("Added", "Contact", None, f"Added contact: {name}")

    def update_contact(self, contact_id, name, phone, email, company_id, notes):
        self.conn.execute(
            "UPDATE contacts SET name=?, phone=?, email=?, company_id=?, notes=? WHERE id=?",
            (name, phone, email, company_id, notes, contact_id)
        )
        self.conn.commit()
        self.log_activity("Updated", "Contact", contact_id, f"Updated contact: {name}")

    def delete_contact(self, contact_id):
        self.conn.execute("DELETE FROM contacts WHERE id=?", (contact_id,))
        self.conn.commit()
        self.log_activity("Deleted", "Contact", contact_id, f"Deleted contact ID: {contact_id}")

    # ========== Companies ==========
    def fetch_all_companies(self):
        cur = self.conn.execute("SELECT * FROM companies ORDER BY name")
        return cur.fetchall()

    def search_companies(self, keyword):
        cur = self.conn.execute(
            "SELECT * FROM companies WHERE name LIKE ? OR phone LIKE ? OR website LIKE ? ORDER BY name",
            (f'%{keyword}%', f'%{keyword}%', f'%{keyword}%')
        )
        return cur.fetchall()

    def add_company(self, name, address, phone, website, notes):
        self.conn.execute(
            "INSERT INTO companies (name, address, phone, website, notes) VALUES (?, ?, ?, ?, ?)",
            (name, address, phone, website, notes)
        )
        self.conn.commit()
        self.log_activity("Added", "Company", None, f"Added company: {name}")

    def update_company(self, company_id, name, address, phone, website, notes):
        self.conn.execute(
            "UPDATE companies SET name=?, address=?, phone=?, website=?, notes=? WHERE id=?",
            (name, address, phone, website, notes, company_id)
        )
        self.conn.commit()
        self.log_activity("Updated", "Company", company_id, f"Updated company: {name}")

    def delete_company(self, company_id):
        self.conn.execute("DELETE FROM companies WHERE id=?", (company_id,))
        self.conn.commit()
        self.log_activity("Deleted", "Company", company_id, f"Deleted company ID: {company_id}")

    def get_company_list(self):
        cur = self.conn.execute("SELECT id, name FROM companies ORDER BY name")
        return cur.fetchall()

    # ========== Attachments ==========
    def fetch_attachments(self, contact_id):
        cur = self.conn.execute("SELECT id, filename FROM attachments WHERE contact_id=?", (contact_id,))
        return cur.fetchall()

    def add_attachment(self, contact_id, filename, data):
        self.conn.execute("INSERT INTO attachments (contact_id, filename, data) VALUES (?, ?, ?)",
                          (contact_id, filename, data))
        self.conn.commit()

    def delete_attachment(self, attachment_id):
        self.conn.execute("DELETE FROM attachments WHERE id=?", (attachment_id,))
        self.conn.commit()

    # ========== Communication Log ==========
    def fetch_communication_log(self, contact_id):
        cur = self.conn.execute(
            "SELECT id, type, date, description FROM communication_log WHERE contact_id=? ORDER BY date DESC",
            (contact_id,))
        return cur.fetchall()

    def add_communication(self, contact_id, comm_type, date, description):
        self.conn.execute(
            "INSERT INTO communication_log (contact_id, type, date, description) VALUES (?, ?, ?, ?)",
            (contact_id, comm_type, date.toString("yyyy-MM-dd") if isinstance(date, QDate) else date, description))
        self.conn.commit()

    # ========== Tasks ==========
    def fetch_all_tasks(self):
        cur = self.conn.execute("""
            SELECT t.id, t.title, t.description, t.due_date, t.status,
                   c.name
            FROM tasks t LEFT JOIN contacts c ON t.contact_id = c.id
            ORDER BY t.due_date
        """)
        return cur.fetchall()

    def search_tasks(self, keyword):
        cur = self.conn.execute("""
            SELECT t.id, t.title, t.description, t.due_date, t.status,
                   c.name
            FROM tasks t LEFT JOIN contacts c ON t.contact_id = c.id
            WHERE t.title LIKE ? OR t.description LIKE ? OR c.name LIKE ?
            ORDER BY t.due_date
        """, (f'%{keyword}%', f'%{keyword}%', f'%{keyword}%'))
        return cur.fetchall()

    def add_task(self, title, description, due_date, status, contact_id):
        self.conn.execute(
            "INSERT INTO tasks (title, description, due_date, status, contact_id) VALUES (?, ?, ?, ?, ?)",
            (title, description, due_date.toString("yyyy-MM-dd") if due_date else None, status, contact_id)
        )
        self.conn.commit()
        self.log_activity("Added", "Task", None, f"Added task: {title}")

    def update_task(self, task_id, title, description, due_date, status, contact_id):
        self.conn.execute(
            "UPDATE tasks SET title=?, description=?, due_date=?, status=?, contact_id=? WHERE id=?",
            (title, description, due_date.toString("yyyy-MM-dd") if due_date else None, status, contact_id, task_id)
        )
        self.conn.commit()
        self.log_activity("Updated", "Task", task_id, f"Updated task: {title}")

    def delete_task(self, task_id):
        self.conn.execute("DELETE FROM tasks WHERE id=?", (task_id,))
        self.conn.commit()
        self.log_activity("Deleted", "Task", task_id, f"Deleted task ID: {task_id}")

    # ========== Deals ==========
    def fetch_all_deals(self):
        cur = self.conn.execute("""
            SELECT d.id, d.title, d.value, d.stage, d.expected_close_date,
                   c.name
            FROM deals d LEFT JOIN contacts c ON d.contact_id = c.id
            ORDER BY d.stage
        """)
        return cur.fetchall()

    def search_deals(self, keyword):
        cur = self.conn.execute("""
            SELECT d.id, d.title, d.value, d.stage, d.expected_close_date,
                   c.name
            FROM deals d LEFT JOIN contacts c ON d.contact_id = c.id
            WHERE d.title LIKE ? OR c.name LIKE ? OR d.stage LIKE ?
            ORDER BY d.stage
        """, (f'%{keyword}%', f'%{keyword}%', f'%{keyword}%'))
        return cur.fetchall()

    def add_deal(self, title, value, stage, expected_close_date, contact_id):
        self.conn.execute(
            "INSERT INTO deals (title, value, stage, expected_close_date, contact_id) VALUES (?, ?, ?, ?, ?)",
            (title, value, stage,
             expected_close_date.toString("yyyy-MM-dd") if expected_close_date else None,
             contact_id)
        )
        self.conn.commit()
        self.log_activity("Added", "Deal", None, f"Added deal: {title}")

    def update_deal(self, deal_id, title, value, stage, expected_close_date, contact_id):
        self.conn.execute(
            "UPDATE deals SET title=?, value=?, stage=?, expected_close_date=?, contact_id=? WHERE id=?",
            (title, value, stage,
             expected_close_date.toString("yyyy-MM-dd") if expected_close_date else None,
             contact_id, deal_id)
        )
        self.conn.commit()
        self.log_activity("Updated", "Deal", deal_id, f"Updated deal: {title}")

    def update_deal_stage(self, deal_id, new_stage):
        self.conn.execute("UPDATE deals SET stage=? WHERE id=?", (new_stage, deal_id))
        self.conn.commit()

    def delete_deal(self, deal_id):
        self.conn.execute("DELETE FROM deals WHERE id=?", (deal_id,))
        self.conn.commit()
        self.log_activity("Deleted", "Deal", deal_id, f"Deleted deal ID: {deal_id}")

    # ========== Helpers ==========
    def get_contact_list(self):
        cur = self.conn.execute("SELECT id, name FROM contacts ORDER BY name")
        return cur.fetchall()

    # ========== Scripts ==========
    def fetch_all_scripts(self):
        cur = self.conn.execute("SELECT id, name, content FROM scripts ORDER BY name")
        return cur.fetchall()

    def add_script(self, name, content):
        self.conn.execute("INSERT INTO scripts (name, content) VALUES (?, ?)", (name, content))
        self.conn.commit()

    def update_script(self, script_id, name, content):
        self.conn.execute("UPDATE scripts SET name=?, content=? WHERE id=?", (name, content, script_id))
        self.conn.commit()

    def delete_script(self, script_id):
        self.conn.execute("DELETE FROM scripts WHERE id=?", (script_id,))
        self.conn.commit()

    # ========== Dashboard Stats ==========
    def total_contacts(self):
        return self.conn.execute("SELECT COUNT(*) FROM contacts").fetchone()[0]

    def open_deals(self):
        return self.conn.execute("SELECT COUNT(*) FROM deals WHERE stage NOT IN ('Won', 'Lost')").fetchone()[0]

    def total_revenue(self):
        cur = self.conn.execute("SELECT SUM(value) FROM deals WHERE stage='Won'")
        val = cur.fetchone()[0]
        return val if val else 0.0

    def close_ratio(self):
        total = self.conn.execute("SELECT COUNT(*) FROM deals").fetchone()[0]
        if total == 0:
            return 0.0
        won = self.conn.execute("SELECT COUNT(*) FROM deals WHERE stage='Won'").fetchone()[0]
        return round(won / total * 100, 1)

    # ========== EXPORT/IMPORT CSV ==========
    def export_contacts_csv(self, filepath):
        contacts = self.fetch_all_contacts()
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(["ID", "Name", "Phone", "Email", "Company", "Notes"])
            for c in contacts:
                writer.writerow(c)

    def import_contacts_csv(self, filepath):
        with open(filepath, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                name = row.get("Name", "").strip()
                if not name:
                    continue
                phone = row.get("Phone", "")
                email = row.get("Email", "")
                company = row.get("Company", "")
                notes = row.get("Notes", "")
                # find company_id if company name given
                company_id = None
                if company:
                    cur = self.conn.execute("SELECT id FROM companies WHERE name=?", (company,))
                    comp = cur.fetchone()
                    if comp:
                        company_id = comp[0]
                self.add_contact(name, phone, email, company_id, notes)

    def export_companies_csv(self, filepath):
        companies = self.fetch_all_companies()
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(["ID", "Name", "Address", "Phone", "Website", "Notes"])
            for c in companies:
                writer.writerow(c)

    def import_companies_csv(self, filepath):
        with open(filepath, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                name = row.get("Name", "").strip()
                if not name:
                    continue
                address = row.get("Address", "")
                phone = row.get("Phone", "")
                website = row.get("Website", "")
                notes = row.get("Notes", "")
                self.add_company(name, address, phone, website, notes)

    def export_tasks_csv(self, filepath):
        tasks = self.fetch_all_tasks()
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(["ID", "Title", "Description", "Due Date", "Status", "Contact"])
            for t in tasks:
                writer.writerow(t)

    def import_tasks_csv(self, filepath):
        with open(filepath, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                title = row.get("Title", "").strip()
                if not title:
                    continue
                description = row.get("Description", "")
                due_date = row.get("Due Date", None)
                status = row.get("Status", "Pending")
                contact_name = row.get("Contact", "")
                contact_id = None
                if contact_name:
                    cur = self.conn.execute("SELECT id FROM contacts WHERE name=?", (contact_name,))
                    c = cur.fetchone()
                    if c:
                        contact_id = c[0]
                self.add_task(title, description, due_date if due_date else None, status, contact_id)

    def export_deals_csv(self, filepath):
        deals = self.fetch_all_deals()
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(["ID", "Title", "Value", "Stage", "Expected Close", "Contact"])
            for d in deals:
                writer.writerow(d)

    def import_deals_csv(self, filepath):
        with open(filepath, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                title = row.get("Title", "").strip()
                if not title:
                    continue
                try:
                    value = float(row.get("Value", 0))
                except:
                    value = 0.0
                stage = row.get("Stage", "Lead")
                expected_close = row.get("Expected Close", None)
                contact_name = row.get("Contact", "")
                contact_id = None
                if contact_name:
                    cur = self.conn.execute("SELECT id FROM contacts WHERE name=?", (contact_name,))
                    c = cur.fetchone()
                    if c:
                        contact_id = c[0]
                self.add_deal(title, value, stage, expected_close if expected_close else None, contact_id)


# --------------------------- TABLE MODELS (unchanged) ---------------------------
class ContactsTableModel(QAbstractTableModel):
    def __init__(self, data=None):
        super().__init__()
        self._data = data or []
        self.headers = ["ID", "Name", "Phone", "Email", "Company", "Notes"]
    def rowCount(self, parent=None): return len(self._data)
    def columnCount(self, parent=None): return len(self.headers)
    def data(self, index, role=Qt.ItemDataRole.DisplayRole):
        if not index.isValid(): return None
        if role == Qt.ItemDataRole.DisplayRole:
            row = self._data[index.row()]
            col = index.column()
            if col == 0: return str(row[0])
            if col == 1: return row[1]
            if col == 2: return row[2] if row[2] else ""
            if col == 3: return row[3] if row[3] else ""
            if col == 4: return row[4] if row[4] else ""
            if col == 5: return row[5] if row[5] else ""
        return None
    def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
        if role == Qt.ItemDataRole.DisplayRole:
            if orientation == Qt.Orientation.Horizontal: return self.headers[section]
            else: return str(section + 1)
        return None
    def refresh_data(self, new_data):
        self.beginResetModel()
        self._data = new_data
        self.endResetModel()

class CompaniesTableModel(QAbstractTableModel):
    def __init__(self, data=None):
        super().__init__()
        self._data = data or []
        self.headers = ["ID", "Name", "Address", "Phone", "Website", "Notes"]
    def rowCount(self, parent=None): return len(self._data)
    def columnCount(self, parent=None): return len(self.headers)
    def data(self, index, role=Qt.ItemDataRole.DisplayRole):
        if not index.isValid(): return None
        if role == Qt.ItemDataRole.DisplayRole:
            row = self._data[index.row()]
            return str(row[index.column()]) if row[index.column()] else ""
        return None
    def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
        if role == Qt.ItemDataRole.DisplayRole:
            if orientation == Qt.Orientation.Horizontal: return self.headers[section]
            else: return str(section + 1)
        return None
    def refresh_data(self, new_data):
        self.beginResetModel()
        self._data = new_data
        self.endResetModel()

class TasksTableModel(QAbstractTableModel):
    def __init__(self, data=None):
        super().__init__()
        self._data = data or []
        self.headers = ["ID", "Title", "Description", "Due Date", "Status", "Contact"]
    def rowCount(self, parent=None): return len(self._data)
    def columnCount(self, parent=None): return len(self.headers)
    def data(self, index, role=Qt.ItemDataRole.DisplayRole):
        if not index.isValid(): return None
        if role == Qt.ItemDataRole.DisplayRole:
            val = self._data[index.row()][index.column()]
            return str(val if val is not None else "")
        return None
    def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
        if role == Qt.ItemDataRole.DisplayRole:
            if orientation == Qt.Orientation.Horizontal: return self.headers[section]
            else: return str(section + 1)
        return None
    def refresh_data(self, new_data):
        self.beginResetModel()
        self._data = new_data
        self.endResetModel()


# --------------------------- CONTACTS TAB (with Export/Import) ---------------------------
class ContactsTab(QWidget):
    def __init__(self, db):
        super().__init__()
        self.db = db
        self.model = ContactsTableModel()
        self.current_id = None

        main_layout = QVBoxLayout(self)

        # Search + Export/Import
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("Search:"))
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search contacts...")
        self.search_input.textChanged.connect(self.on_search)
        search_layout.addWidget(self.search_input)

        self.export_btn = QPushButton("Export CSV")
        self.export_btn.clicked.connect(self.export_data)
        self.import_btn = QPushButton("Import CSV")
        self.import_btn.clicked.connect(self.import_data)
        search_layout.addWidget(self.export_btn)
        search_layout.addWidget(self.import_btn)
        main_layout.addLayout(search_layout)

        content = QHBoxLayout()
        self.table = QTableView()
        self.table.setModel(self.model)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.clicked.connect(self.on_contact_select)
        self.table.setColumnHidden(0, True)
        content.addWidget(self.table, 2)

        right_panel = QVBoxLayout()
        right_panel.setContentsMargins(0, 0, 0, 0)
        details = QTabWidget()
        details.setTabPosition(QTabWidget.TabPosition.North)

        # Basic Info
        basic_widget = QWidget()
        basic_main_layout = QVBoxLayout()
        group_box = QGroupBox("Contact Details")
        form_layout = QFormLayout()
        self.name_input = QLineEdit()
        self.phone_input = QLineEdit()
        self.email_input = QLineEdit()
        self.company_combo = QComboBox()
        self.company_combo.addItem("-- None --", None)
        self.notes_input = QLineEdit()
        form_layout.addRow("Name *:", self.name_input)
        form_layout.addRow("Phone:", self.phone_input)
        form_layout.addRow("Email:", self.email_input)
        form_layout.addRow("Company:", self.company_combo)
        form_layout.addRow("Notes:", self.notes_input)
        group_box.setLayout(form_layout)
        basic_main_layout.addWidget(group_box)

        btn_layout = QVBoxLayout()
        self.add_btn = QPushButton("Add")
        self.edit_btn = QPushButton("Edit")
        self.delete_btn = QPushButton("Delete")
        self.clear_btn = QPushButton("Clear")
        for btn in (self.add_btn, self.edit_btn, self.delete_btn, self.clear_btn):
            btn.setMinimumHeight(40)
            btn.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
            btn_layout.addWidget(btn)
        basic_main_layout.addLayout(btn_layout)
        basic_widget.setLayout(basic_main_layout)
        details.addTab(basic_widget, "Basic Info")

        # Communication Log
        comm_widget = QWidget()
        comm_layout = QVBoxLayout()
        self.comm_list = QListWidget()
        self.comm_type = QComboBox()
        self.comm_type.addItems(["Call", "Meeting", "Email"])
        self.comm_date = QDateEdit(QDate.currentDate())
        self.comm_desc = QLineEdit()
        comm_add_btn = QPushButton("Add Log Entry")
        comm_add_btn.clicked.connect(self.add_comm_entry)
        comm_layout.addWidget(QLabel("Type:"))
        comm_layout.addWidget(self.comm_type)
        comm_layout.addWidget(QLabel("Date:"))
        comm_layout.addWidget(self.comm_date)
        comm_layout.addWidget(QLabel("Description:"))
        comm_layout.addWidget(self.comm_desc)
        comm_layout.addWidget(comm_add_btn)
        comm_layout.addWidget(self.comm_list)
        comm_widget.setLayout(comm_layout)
        details.addTab(comm_widget, "Communication Log")

        # Attachments
        attach_widget = QWidget()
        attach_layout = QVBoxLayout()
        self.attach_list = QListWidget()
        attach_add_btn = QPushButton("Add Attachment")
        attach_add_btn.clicked.connect(self.add_attachment)
        attach_remove_btn = QPushButton("Remove Selected")
        attach_remove_btn.clicked.connect(self.remove_attachment)
        attach_layout.addWidget(attach_add_btn)
        attach_layout.addWidget(attach_remove_btn)
        attach_layout.addWidget(self.attach_list)
        attach_widget.setLayout(attach_layout)
        details.addTab(attach_widget, "Attachments")

        right_panel.addWidget(details)
        content.addLayout(right_panel, 1)
        main_layout.addLayout(content)

        self.add_btn.clicked.connect(self.add_contact)
        self.edit_btn.clicked.connect(self.edit_contact)
        self.delete_btn.clicked.connect(self.delete_contact)
        self.clear_btn.clicked.connect(self.clear_form)

        self.load_companies()
        self.load_data()

    def load_companies(self):
        self.company_combo.clear()
        self.company_combo.addItem("-- None --", None)
        for cid, name in self.db.get_company_list():
            self.company_combo.addItem(name, cid)

    def load_data(self):
        data = self.db.fetch_all_contacts()
        self.model.refresh_data(data)

    def on_search(self, text):
        data = self.db.search_contacts(text) if text else self.db.fetch_all_contacts()
        self.model.refresh_data(data)

    def export_data(self):
        filepath, _ = QFileDialog.getSaveFileName(self, "Export Contacts", "", "CSV Files (*.csv);;Excel Files (*.xlsx)")
        if not filepath:
            return
        if filepath.endswith('.xlsx'):
            try:
                import openpyxl
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Contacts"
                contacts = self.db.fetch_all_contacts()
                ws.append(["ID", "Name", "Phone", "Email", "Company", "Notes"])
                for c in contacts:
                    ws.append(list(c))
                wb.save(filepath)
                QMessageBox.information(self, "Success", "Contacts exported to Excel.")
            except ImportError:
                QMessageBox.warning(self, "Error", "openpyxl library not installed. Saving as CSV.")
                self.db.export_contacts_csv(filepath)
                QMessageBox.information(self, "Success", "Contacts exported to CSV.")
        else:
            self.db.export_contacts_csv(filepath)
            QMessageBox.information(self, "Success", "Contacts exported to CSV.")

    def import_data(self):
        filepath, _ = QFileDialog.getOpenFileName(self, "Import Contacts", "", "CSV Files (*.csv);;Excel Files (*.xlsx)")
        if not filepath:
            return
        if filepath.endswith('.xlsx'):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(filepath)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    data = dict(zip(headers, row))
                    name = data.get("Name", "").strip()
                    if not name:
                        continue
                    phone = data.get("Phone", "")
                    email = data.get("Email", "")
                    company = data.get("Company", "")
                    notes = data.get("Notes", "")
                    company_id = None
                    if company:
                        cur = self.db.conn.execute("SELECT id FROM companies WHERE name=?", (company,))
                        comp = cur.fetchone()
                        if comp:
                            company_id = comp[0]
                    self.db.add_contact(name, phone, email, company_id, notes)
                self.load_data()
                QMessageBox.information(self, "Success", "Contacts imported from Excel.")
            except ImportError:
                QMessageBox.warning(self, "Error", "openpyxl library not installed.")
                return
        else:
            self.db.import_contacts_csv(filepath)
            self.load_data()
            QMessageBox.information(self, "Success", "Contacts imported from CSV.")

    def on_contact_select(self, index):
        row = index.row()
        record = self.model._data[row]
        self.current_id = record[0]
        self.name_input.setText(record[1] or "")
        self.phone_input.setText(record[2] or "")
        self.email_input.setText(record[3] or "")
        company_name = record[4]
        idx = self.company_combo.findText(company_name if company_name else "-- None --")
        if idx >= 0:
            self.company_combo.setCurrentIndex(idx)
        else:
            self.company_combo.setCurrentIndex(0)
        self.notes_input.setText(record[5] or "")
        self.load_communication_log()
        self.load_attachments()

    def load_communication_log(self):
        self.comm_list.clear()
        if not self.current_id:
            return
        for eid, etype, edate, edesc in self.db.fetch_communication_log(self.current_id):
            self.comm_list.addItem(f"{edate} - {etype}: {edesc}")

    def add_comm_entry(self):
        if not self.current_id:
            QMessageBox.warning(self, "Warning", "Select a contact first.")
            return
        self.db.add_communication(self.current_id, self.comm_type.currentText(),
                                  self.comm_date.date(), self.comm_desc.text().strip())
        self.comm_desc.clear()
        self.load_communication_log()

    def load_attachments(self):
        self.attach_list.clear()
        if not self.current_id:
            return
        for aid, fname in self.db.fetch_attachments(self.current_id):
            item = QListWidgetItem(fname)
            item.setData(Qt.ItemDataRole.UserRole, aid)
            self.attach_list.addItem(item)

    def add_attachment(self):
        if not self.current_id:
            QMessageBox.warning(self, "Warning", "Select a contact first.")
            return
        file_path, _ = QFileDialog.getOpenFileName(self, "Select file")
        if file_path:
            filename = os.path.basename(file_path)
            with open(file_path, 'rb') as f:
                data = f.read()
            self.db.add_attachment(self.current_id, filename, data)
            self.load_attachments()

    def remove_attachment(self):
        item = self.attach_list.currentItem()
        if not item:
            return
        aid = item.data(Qt.ItemDataRole.UserRole)
        self.db.delete_attachment(aid)
        self.load_attachments()

    def clear_form(self):
        self.current_id = None
        for inp in [self.name_input, self.phone_input, self.email_input, self.notes_input]:
            inp.clear()
        self.company_combo.setCurrentIndex(0)
        self.comm_list.clear()
        self.attach_list.clear()

    def add_contact(self):
        name = self.name_input.text().strip()
        if not name:
            QMessageBox.warning(self, "Warning", "Name is required.")
            return
        self.db.add_contact(name, self.phone_input.text().strip(), self.email_input.text().strip(),
                            self.company_combo.currentData(), self.notes_input.text().strip())
        self.clear_form()
        self.load_data()
        QMessageBox.information(self, "Success", "Contact added.")

    def edit_contact(self):
        if not self.current_id:
            QMessageBox.warning(self, "Warning", "Select a contact first.")
            return
        name = self.name_input.text().strip()
        if not name:
            QMessageBox.warning(self, "Warning", "Name is required.")
            return
        self.db.update_contact(self.current_id, name, self.phone_input.text().strip(),
                               self.email_input.text().strip(), self.company_combo.currentData(),
                               self.notes_input.text().strip())
        self.clear_form()
        self.load_data()
        QMessageBox.information(self, "Success", "Contact updated.")

    def delete_contact(self):
        if not self.current_id:
            QMessageBox.warning(self, "Warning", "Select a contact first.")
            return
        reply = QMessageBox.question(self, "Confirm", "Delete this contact?",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            self.db.delete_contact(self.current_id)
            self.clear_form()
            self.load_data()
            QMessageBox.information(self, "Success", "Contact deleted.")


# --------------------------- COMPANIES TAB (with Export/Import) ---------------------------
class CompaniesTab(QWidget):
    def __init__(self, db):
        super().__init__()
        self.db = db
        self.model = CompaniesTableModel()
        self.current_id = None

        main_layout = QVBoxLayout(self)

        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("Search:"))
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search companies...")
        self.search_input.textChanged.connect(self.on_search)
        search_layout.addWidget(self.search_input)

        self.export_btn = QPushButton("Export CSV")
        self.export_btn.clicked.connect(self.export_data)
        self.import_btn = QPushButton("Import CSV")
        self.import_btn.clicked.connect(self.import_data)
        search_layout.addWidget(self.export_btn)
        search_layout.addWidget(self.import_btn)
        main_layout.addLayout(search_layout)

        content = QHBoxLayout()
        self.table = QTableView()
        self.table.setModel(self.model)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.clicked.connect(self.on_row_select)
        self.table.setColumnHidden(0, True)
        content.addWidget(self.table, 2)

        form_container = QWidget()
        form_layout = QVBoxLayout(form_container)
        form_group = QGroupBox("Company Details")
        inner_form = QFormLayout()
        self.name_input = QLineEdit()
        self.address_input = QLineEdit()
        self.phone_input = QLineEdit()
        self.website_input = QLineEdit()
        self.notes_input = QLineEdit()
        inner_form.addRow("Name *:", self.name_input)
        inner_form.addRow("Address:", self.address_input)
        inner_form.addRow("Phone:", self.phone_input)
        inner_form.addRow("Website:", self.website_input)
        inner_form.addRow("Notes:", self.notes_input)
        form_group.setLayout(inner_form)
        form_layout.addWidget(form_group)

        btn_layout = QVBoxLayout()
        add_btn = QPushButton("Add")
        edit_btn = QPushButton("Edit")
        delete_btn = QPushButton("Delete")
        clear_btn = QPushButton("Clear")
        for btn in (add_btn, edit_btn, delete_btn, clear_btn):
            btn.setMinimumHeight(40)
            btn.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        add_btn.clicked.connect(self.add_company)
        edit_btn.clicked.connect(self.edit_company)
        delete_btn.clicked.connect(self.delete_company)
        clear_btn.clicked.connect(self.clear_form)
        btn_layout.addWidget(add_btn)
        btn_layout.addWidget(edit_btn)
        btn_layout.addWidget(delete_btn)
        btn_layout.addWidget(clear_btn)
        form_layout.addLayout(btn_layout)

        content.addWidget(form_container, 1)
        main_layout.addLayout(content)

        self.load_data()

    def load_data(self):
        data = self.db.fetch_all_companies()
        self.model.refresh_data(data)

    def on_search(self, text):
        data = self.db.search_companies(text) if text else self.db.fetch_all_companies()
        self.model.refresh_data(data)

    def export_data(self):
        filepath, _ = QFileDialog.getSaveFileName(self, "Export Companies", "", "CSV Files (*.csv);;Excel Files (*.xlsx)")
        if not filepath:
            return
        if filepath.endswith('.xlsx'):
            try:
                import openpyxl
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Companies"
                companies = self.db.fetch_all_companies()
                ws.append(["ID", "Name", "Address", "Phone", "Website", "Notes"])
                for c in companies:
                    ws.append(list(c))
                wb.save(filepath)
                QMessageBox.information(self, "Success", "Companies exported to Excel.")
            except ImportError:
                QMessageBox.warning(self, "Error", "openpyxl library not installed. Saving as CSV.")
                self.db.export_companies_csv(filepath)
                QMessageBox.information(self, "Success", "Companies exported to CSV.")
        else:
            self.db.export_companies_csv(filepath)
            QMessageBox.information(self, "Success", "Companies exported to CSV.")

    def import_data(self):
        filepath, _ = QFileDialog.getOpenFileName(self, "Import Companies", "", "CSV Files (*.csv);;Excel Files (*.xlsx)")
        if not filepath:
            return
        if filepath.endswith('.xlsx'):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(filepath)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    data = dict(zip(headers, row))
                    name = data.get("Name", "").strip()
                    if not name:
                        continue
                    address = data.get("Address", "")
                    phone = data.get("Phone", "")
                    website = data.get("Website", "")
                    notes = data.get("Notes", "")
                    self.db.add_company(name, address, phone, website, notes)
                self.load_data()
                QMessageBox.information(self, "Success", "Companies imported from Excel.")
            except ImportError:
                QMessageBox.warning(self, "Error", "openpyxl library not installed.")
                return
        else:
            self.db.import_companies_csv(filepath)
            self.load_data()
            QMessageBox.information(self, "Success", "Companies imported from CSV.")

    def on_row_select(self, index):
        row = index.row()
        record = self.model._data[row]
        self.current_id = record[0]
        self.name_input.setText(record[1] or "")
        self.address_input.setText(record[2] or "")
        self.phone_input.setText(record[3] or "")
        self.website_input.setText(record[4] or "")
        self.notes_input.setText(record[5] or "")

    def clear_form(self):
        self.current_id = None
        for inp in (self.name_input, self.address_input, self.phone_input, self.website_input, self.notes_input):
            inp.clear()

    def add_company(self):
        name = self.name_input.text().strip()
        if not name:
            QMessageBox.warning(self, "Warning", "Name is required.")
            return
        self.db.add_company(name, self.address_input.text().strip(), self.phone_input.text().strip(),
                            self.website_input.text().strip(), self.notes_input.text().strip())
        self.clear_form()
        self.load_data()
        QMessageBox.information(self, "Success", "Company added.")

    def edit_company(self):
        if not self.current_id:
            QMessageBox.warning(self, "Warning", "Select a company first.")
            return
        name = self.name_input.text().strip()
        if not name:
            QMessageBox.warning(self, "Warning", "Name is required.")
            return
        self.db.update_company(self.current_id, name, self.address_input.text().strip(),
                               self.phone_input.text().strip(), self.website_input.text().strip(),
                               self.notes_input.text().strip())
        self.clear_form()
        self.load_data()
        QMessageBox.information(self, "Success", "Company updated.")

    def delete_company(self):
        if not self.current_id:
            QMessageBox.warning(self, "Warning", "Select a company first.")
            return
        reply = QMessageBox.question(self, "Confirm", "Delete this company?",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            self.db.delete_company(self.current_id)
            self.clear_form()
            self.load_data()
            QMessageBox.information(self, "Success", "Company deleted.")


# --------------------------- TASKS TAB (with Export/Import) ---------------------------
class TasksTab(QWidget):
    def __init__(self, db):
        super().__init__()
        self.db = db
        self.model = TasksTableModel()
        self.current_id = None

        main_layout = QVBoxLayout(self)

        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("Search:"))
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search tasks...")
        self.search_input.textChanged.connect(self.on_search)
        search_layout.addWidget(self.search_input)

        self.export_btn = QPushButton("Export CSV")
        self.export_btn.clicked.connect(self.export_data)
        self.import_btn = QPushButton("Import CSV")
        self.import_btn.clicked.connect(self.import_data)
        search_layout.addWidget(self.export_btn)
        search_layout.addWidget(self.import_btn)
        main_layout.addLayout(search_layout)

        content = QHBoxLayout()
        self.table = QTableView()
        self.table.setModel(self.model)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.clicked.connect(self.on_row_select)
        self.table.setColumnHidden(0, True)
        content.addWidget(self.table, 2)

        form_container = QWidget()
        form_layout = QVBoxLayout(form_container)
        form_group = QGroupBox("Task Details")
        inner_form = QFormLayout()
        self.title_input = QLineEdit()
        self.desc_input = QLineEdit()
        self.due_date = QDateEdit(QDate.currentDate())
        self.status_combo = QComboBox()
        self.status_combo.addItems(["Pending", "In Progress", "Completed"])
        self.contact_combo = QComboBox()
        self.contact_combo.addItem("-- None --", None)
        inner_form.addRow("Title *:", self.title_input)
        inner_form.addRow("Description:", self.desc_input)
        inner_form.addRow("Due Date:", self.due_date)
        inner_form.addRow("Status:", self.status_combo)
        inner_form.addRow("Contact:", self.contact_combo)
        form_group.setLayout(inner_form)
        form_layout.addWidget(form_group)

        btn_layout = QVBoxLayout()
        add_btn = QPushButton("Add")
        edit_btn = QPushButton("Edit")
        delete_btn = QPushButton("Delete")
        clear_btn = QPushButton("Clear")
        for btn in (add_btn, edit_btn, delete_btn, clear_btn):
            btn.setMinimumHeight(40)
            btn.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        add_btn.clicked.connect(self.add_task)
        edit_btn.clicked.connect(self.edit_task)
        delete_btn.clicked.connect(self.delete_task)
        clear_btn.clicked.connect(self.clear_form)
        btn_layout.addWidget(add_btn)
        btn_layout.addWidget(edit_btn)
        btn_layout.addWidget(delete_btn)
        btn_layout.addWidget(clear_btn)
        form_layout.addLayout(btn_layout)

        content.addWidget(form_container, 1)
        main_layout.addLayout(content)

        self.load_contacts()
        self.load_data()

    def load_contacts(self):
        self.contact_combo.clear()
        self.contact_combo.addItem("-- None --", None)
        for cid, name in self.db.get_contact_list():
            self.contact_combo.addItem(name, cid)

    def load_data(self):
        data = self.db.fetch_all_tasks()
        self.model.refresh_data(data)

    def on_search(self, text):
        data = self.db.search_tasks(text) if text else self.db.fetch_all_tasks()
        self.model.refresh_data(data)

    def export_data(self):
        filepath, _ = QFileDialog.getSaveFileName(self, "Export Tasks", "", "CSV Files (*.csv);;Excel Files (*.xlsx)")
        if not filepath:
            return
        if filepath.endswith('.xlsx'):
            try:
                import openpyxl
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Tasks"
                tasks = self.db.fetch_all_tasks()
                ws.append(["ID", "Title", "Description", "Due Date", "Status", "Contact"])
                for t in tasks:
                    ws.append(list(t))
                wb.save(filepath)
                QMessageBox.information(self, "Success", "Tasks exported to Excel.")
            except ImportError:
                QMessageBox.warning(self, "Error", "openpyxl library not installed. Saving as CSV.")
                self.db.export_tasks_csv(filepath)
                QMessageBox.information(self, "Success", "Tasks exported to CSV.")
        else:
            self.db.export_tasks_csv(filepath)
            QMessageBox.information(self, "Success", "Tasks exported to CSV.")

    def import_data(self):
        filepath, _ = QFileDialog.getOpenFileName(self, "Import Tasks", "", "CSV Files (*.csv);;Excel Files (*.xlsx)")
        if not filepath:
            return
        if filepath.endswith('.xlsx'):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(filepath)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    data = dict(zip(headers, row))
                    title = data.get("Title", "").strip()
                    if not title:
                        continue
                    description = data.get("Description", "")
                    due_date = data.get("Due Date", None)
                    status = data.get("Status", "Pending")
                    contact_name = data.get("Contact", "")
                    contact_id = None
                    if contact_name:
                        cur = self.db.conn.execute("SELECT id FROM contacts WHERE name=?", (contact_name,))
                        c = cur.fetchone()
                        if c:
                            contact_id = c[0]
                    self.db.add_task(title, description, due_date if due_date else None, status, contact_id)
                self.load_data()
                QMessageBox.information(self, "Success", "Tasks imported from Excel.")
            except ImportError:
                QMessageBox.warning(self, "Error", "openpyxl library not installed.")
                return
        else:
            self.db.import_tasks_csv(filepath)
            self.load_data()
            QMessageBox.information(self, "Success", "Tasks imported from CSV.")

    def on_row_select(self, index):
        row = index.row()
        record = self.model._data[row]
        self.current_id = record[0]
        self.title_input.setText(record[1] or "")
        self.desc_input.setText(record[2] or "")
        if record[3]:
            self.due_date.setDate(QDate.fromString(record[3], "yyyy-MM-dd"))
        else:
            self.due_date.setDate(QDate.currentDate())
        idx = self.status_combo.findText(record[4])
        if idx >= 0:
            self.status_combo.setCurrentIndex(idx)
        contact_name = record[5] if record[5] else "-- None --"
        contact_idx = self.contact_combo.findText(contact_name)
        if contact_idx >= 0:
            self.contact_combo.setCurrentIndex(contact_idx)
        else:
            self.contact_combo.setCurrentIndex(0)

    def clear_form(self):
        self.current_id = None
        self.title_input.clear()
        self.desc_input.clear()
        self.due_date.setDate(QDate.currentDate())
        self.status_combo.setCurrentIndex(0)
        self.contact_combo.setCurrentIndex(0)

    def add_task(self):
        title = self.title_input.text().strip()
        if not title:
            QMessageBox.warning(self, "Warning", "Title is required.")
            return
        self.db.add_task(title, self.desc_input.text().strip(), self.due_date.date(),
                         self.status_combo.currentText(), self.contact_combo.currentData())
        self.clear_form()
        self.load_data()
        QMessageBox.information(self, "Success", "Task added.")

    def edit_task(self):
        if not self.current_id:
            QMessageBox.warning(self, "Warning", "Select a task first.")
            return
        title = self.title_input.text().strip()
        if not title:
            QMessageBox.warning(self, "Warning", "Title is required.")
            return
        self.db.update_task(self.current_id, title, self.desc_input.text().strip(),
                            self.due_date.date(), self.status_combo.currentText(),
                            self.contact_combo.currentData())
        self.clear_form()
        self.load_data()
        QMessageBox.information(self, "Success", "Task updated.")

    def delete_task(self):
        if not self.current_id:
            QMessageBox.warning(self, "Warning", "Select a task first.")
            return
        reply = QMessageBox.question(self, "Confirm", "Delete this task?",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            self.db.delete_task(self.current_id)
            self.clear_form()
            self.load_data()
            QMessageBox.information(self, "Success", "Task deleted.")


# --------------------------- KANBAN DEALS TAB (with Export/Import) ---------------------------
class KanbanDealsTab(QWidget):
    def __init__(self, db):
        super().__init__()
        self.db = db
        self.stages = ["Lead", "Qualified", "Proposal", "Negotiation", "Won", "Lost"]
        self.columns = {}

        main_layout = QVBoxLayout(self)

        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("Search:"))
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search deals...")
        search_layout.addWidget(self.search_input)

        self.export_btn = QPushButton("Export CSV")
        self.export_btn.clicked.connect(self.export_data)
        self.import_btn = QPushButton("Import CSV")
        self.import_btn.clicked.connect(self.import_data)
        search_layout.addWidget(self.export_btn)
        search_layout.addWidget(self.import_btn)
        main_layout.addLayout(search_layout)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll_widget = QWidget()
        grid = QGridLayout(scroll_widget)
        for i, stage in enumerate(self.stages):
            column = KanbanColumn(stage)
            lbl = QLabel(f"<h3>{stage}</h3>")
            lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            grid.addWidget(lbl, 0, i)
            grid.addWidget(column, 1, i)
            self.columns[stage] = column
        scroll.setWidget(scroll_widget)
        main_layout.addWidget(scroll)

        btn_layout = QHBoxLayout()
        add_btn = QPushButton("Add Deal")
        add_btn.clicked.connect(self.add_deal_dialog)
        btn_layout.addWidget(add_btn)
        main_layout.addLayout(btn_layout)

        self.load_pipeline()

    def load_pipeline(self):
        deals = self.db.fetch_all_deals()
        for col in self.columns.values():
            col.clear()
        for deal in deals:
            stage = deal[3]
            if stage in self.columns:
                self.columns[stage].add_deal_card(deal)

    def refresh_pipeline(self):
        for stage, column in self.columns.items():
            for i in range(column.count()):
                item = column.item(i)
                if hasattr(item, 'deal_id'):
                    self.db.update_deal_stage(item.deal_id, stage)
        self.load_pipeline()

    def add_deal_dialog(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("New Deal")
        layout = QFormLayout(dialog)
        title_edit = QLineEdit()
        value_edit = QLineEdit("0.00")
        stage_combo = QComboBox()
        stage_combo.addItems(self.stages)
        close_date = QDateEdit(QDate.currentDate())
        contact_combo = QComboBox()
        contact_combo.addItem("-- None --", None)
        for cid, name in self.db.get_contact_list():
            contact_combo.addItem(name, cid)
        layout.addRow("Title:", title_edit)
        layout.addRow("Value:", value_edit)
        layout.addRow("Stage:", stage_combo)
        layout.addRow("Expected Close:", close_date)
        layout.addRow("Contact:", contact_combo)
        buttons = QHBoxLayout()
        ok_btn = QPushButton("OK")
        cancel_btn = QPushButton("Cancel")
        buttons.addWidget(ok_btn)
        buttons.addWidget(cancel_btn)
        layout.addRow(buttons)
        ok_btn.clicked.connect(dialog.accept)
        cancel_btn.clicked.connect(dialog.reject)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            try:
                value = float(value_edit.text())
            except:
                value = 0.0
            self.db.add_deal(title_edit.text(), value, stage_combo.currentText(),
                             close_date.date(), contact_combo.currentData())
            self.load_pipeline()

    def export_data(self):
        filepath, _ = QFileDialog.getSaveFileName(self, "Export Deals", "", "CSV Files (*.csv);;Excel Files (*.xlsx)")
        if not filepath:
            return
        if filepath.endswith('.xlsx'):
            try:
                import openpyxl
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Deals"
                deals = self.db.fetch_all_deals()
                ws.append(["ID", "Title", "Value", "Stage", "Expected Close", "Contact"])
                for d in deals:
                    ws.append(list(d))
                wb.save(filepath)
                QMessageBox.information(self, "Success", "Deals exported to Excel.")
            except ImportError:
                QMessageBox.warning(self, "Error", "openpyxl library not installed. Saving as CSV.")
                self.db.export_deals_csv(filepath)
                QMessageBox.information(self, "Success", "Deals exported to CSV.")
        else:
            self.db.export_deals_csv(filepath)
            QMessageBox.information(self, "Success", "Deals exported to CSV.")

    def import_data(self):
        filepath, _ = QFileDialog.getOpenFileName(self, "Import Deals", "", "CSV Files (*.csv);;Excel Files (*.xlsx)")
        if not filepath:
            return
        if filepath.endswith('.xlsx'):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(filepath)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    data = dict(zip(headers, row))
                    title = data.get("Title", "").strip()
                    if not title:
                        continue
                    try:
                        value = float(data.get("Value", 0))
                    except:
                        value = 0.0
                    stage = data.get("Stage", "Lead")
                    expected_close = data.get("Expected Close", None)
                    contact_name = data.get("Contact", "")
                    contact_id = None
                    if contact_name:
                        cur = self.db.conn.execute("SELECT id FROM contacts WHERE name=?", (contact_name,))
                        c = cur.fetchone()
                        if c:
                            contact_id = c[0]
                    self.db.add_deal(title, value, stage, expected_close if expected_close else None, contact_id)
                self.load_pipeline()
                QMessageBox.information(self, "Success", "Deals imported from Excel.")
            except ImportError:
                QMessageBox.warning(self, "Error", "openpyxl library not installed.")
                return
        else:
            self.db.import_deals_csv(filepath)
            self.load_pipeline()
            QMessageBox.information(self, "Success", "Deals imported from CSV.")


# --------------------------- DealCard / KanbanColumn (unchanged) ---------------------------
class DealCard(QFrame):
    def __init__(self, deal_id, title, value, close_date, contact_name, parent=None):
        super().__init__(parent)
        self.deal_id = deal_id
        self.setFrameStyle(QFrame.Shape.Box | QFrame.Shadow.Raised)
        self.setMinimumHeight(60)
        layout = QVBoxLayout()
        layout.addWidget(QLabel(f"<b>{title}</b>"))
        layout.addWidget(QLabel(f"${value:,.2f}"))
        if close_date:
            layout.addWidget(QLabel(f"Close: {close_date}"))
        if contact_name:
            layout.addWidget(QLabel(f"Contact: {contact_name}"))
        self.setLayout(layout)

class KanbanColumn(QListWidget):
    def __init__(self, stage, parent=None):
        super().__init__(parent)
        self.stage = stage
        self.setDragDropMode(QAbstractItemView.DragDropMode.DragDrop)
        self.setDefaultDropAction(Qt.DropAction.MoveAction)
        self.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)

    def dropEvent(self, event):
        super().dropEvent(event)
        QTimer.singleShot(100, self.parent().parent().refresh_pipeline)

    def add_deal_card(self, deal):
        item = QListWidgetItem()
        card = DealCard(deal[0], deal[1], deal[2], deal[4], deal[5])
        item.setSizeHint(card.sizeHint())
        self.addItem(item)
        self.setItemWidget(item, card)
        item.deal_id = deal[0]


# --------------------------- CALENDAR TAB ---------------------------
class CalendarTab(QWidget):
    def __init__(self, db):
        super().__init__()
        self.db = db
        layout = QHBoxLayout(self)
        self.calendar = QCalendarWidget()
        self.calendar.setLocale(QLocale(QLocale.Language.English))
        self.calendar.clicked.connect(self.on_date_selected)
        layout.addWidget(self.calendar, 1)

        right_panel = QVBoxLayout()
        self.event_list = QListWidget()
        right_panel.addWidget(QLabel("Tasks/Events on selected date:"))
        right_panel.addWidget(self.event_list)
        layout.addLayout(right_panel, 1)
        self.on_date_selected(self.calendar.selectedDate())

    def on_date_selected(self, date):
        self.event_list.clear()
        date_str = date.toString("yyyy-MM-dd")
        cur = self.db.conn.execute("SELECT title, description, status FROM tasks WHERE due_date=?", (date_str,))
        for title, desc, status in cur.fetchall():
            self.event_list.addItem(f"Task: {title} ({status}) - {desc if desc else ''}")


# --------------------------- DASHBOARD TAB ---------------------------
class DashboardTab(QWidget):
    def __init__(self, db):
        super().__init__()
        self.db = db
        layout = QVBoxLayout(self)

        title = QLabel("<h2>Dashboard</h2>")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title)

        cards_layout = QHBoxLayout()
        self.lbl_contacts = QLabel()
        self.lbl_open_deals = QLabel()
        self.lbl_revenue = QLabel()
        self.lbl_close_ratio = QLabel()

        for lbl, caption in [(self.lbl_contacts, "Total Contacts"),
                             (self.lbl_open_deals, "Open Deals"),
                             (self.lbl_revenue, "Revenue (Won)"),
                             (self.lbl_close_ratio, "Close Ratio")]:
            container = QFrame()
            container.setFrameStyle(QFrame.Shape.Box | QFrame.Shadow.Raised)
            inner = QVBoxLayout()
            inner.addWidget(QLabel(f"<b>{caption}</b>"))
            lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lbl.setFont(QFont("Segoe UI", 16))
            inner.addWidget(lbl)
            container.setLayout(inner)
            cards_layout.addWidget(container)

        layout.addLayout(cards_layout)

        layout.addWidget(QLabel("<b>Recent Activity</b>"))
        self.activity_list = QListWidget()
        layout.addWidget(self.activity_list)

        self.refresh()

    def refresh(self):
        self.lbl_contacts.setText(str(self.db.total_contacts()))
        self.lbl_open_deals.setText(str(self.db.open_deals()))
        self.lbl_revenue.setText(f"${self.db.total_revenue():,.2f}")
        self.lbl_close_ratio.setText(f"{self.db.close_ratio()}%")

        self.activity_list.clear()
        for timestamp, action, entity, details in self.db.fetch_recent_activities(20):
            self.activity_list.addItem(f"[{timestamp}] {action} {entity}: {details}")


# --------------------------- REMINDER ---------------------------
def check_upcoming_tasks(db):
    tomorrow = (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d")
    cur = db.conn.execute("SELECT title, due_date FROM tasks WHERE due_date <= ? AND status != 'Completed'", (tomorrow,))
    tasks = cur.fetchall()
    if tasks:
        msg = "Upcoming tasks:\n" + "\n".join([f"- {t[0]} (due {t[1]})" for t in tasks])
        QMessageBox.information(None, "Task Reminder", msg)


# --------------------------- LUA EDITOR (unchanged) ---------------------------
class LuaHighlighter(QSyntaxHighlighter):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.keywords = ["and","break","do","else","elseif","end","false","for","function","goto","if","in","local","nil","not","or","repeat","return","then","true","until","while"]
        self.keyword_format = QTextCharFormat(); self.keyword_format.setForeground(QColor("#2B579A")); self.keyword_format.setFontWeight(QFont.Weight.Bold)
        self.comment_format = QTextCharFormat(); self.comment_format.setForeground(QColor("#5A8A5A")); self.comment_format.setFontItalic(True)
        self.string_format = QTextCharFormat(); self.string_format.setForeground(QColor("#B05A5A"))
        self.number_format = QTextCharFormat(); self.number_format.setForeground(QColor("#4A7A8A"))
    def highlightBlock(self, text):
        for kw in self.keywords:
            for m in re.finditer(r'\b'+kw+r'\b', text): self.setFormat(m.start(), m.end()-m.start(), self.keyword_format)
        for m in re.finditer(r'--[^\n]*', text): self.setFormat(m.start(), m.end()-m.start(), self.comment_format)
        for m in re.finditer(r'"[^"\\]*(\\.[^"\\]*)*"|\'[^\'\\]*(\\.[^\'\\]*)*\'', text): self.setFormat(m.start(), m.end()-m.start(), self.string_format)
        for m in re.finditer(r'\b\d+(\.\d+)?\b', text): self.setFormat(m.start(), m.end()-m.start(), self.number_format)

class LuaCodeEditor(QTextEdit):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.indent_size = 4
        self.highlighter = LuaHighlighter(self.document())
    def keyPressEvent(self, event):
        if event.key() in (Qt.Key.Key_Return, Qt.Key.Key_Enter): self.auto_indent()
        else: super().keyPressEvent(event)
    def auto_indent(self):
        cursor = self.textCursor(); block = cursor.block(); text = block.text()
        cur_indent = len(text) - len(text.lstrip())
        trimmed = text.strip()
        outdent_kw = ["end","until","else","elseif"]
        need_out = False
        if trimmed:
            if trimmed.split()[0] in outdent_kw: need_out = True
        if need_out and cur_indent >= self.indent_size:
            cursor.movePosition(QTextCursor.MoveOperation.StartOfBlock)
            cursor.movePosition(QTextCursor.MoveOperation.Right, QTextCursor.MoveMode.KeepAnchor, self.indent_size)
            cursor.removeSelectedText(); cur_indent -= self.indent_size
        cursor.movePosition(QTextCursor.MoveOperation.EndOfBlock); cursor.insertText("\n")
        new_indent = cur_indent
        if trimmed:
            if not need_out:
                first_word = trimmed.split()[0]
                if first_word in ["function","if","for","while","repeat"]: new_indent += self.indent_size
                elif trimmed.endswith(("then","do","else")): new_indent += self.indent_size
        cursor.insertText(" " * new_indent); self.setTextCursor(cursor)

class ScriptEditorWindow(QWidget):
    def __init__(self, db):
        super().__init__()
        self.db = db
        self.setWindowTitle("Lua Script Editor - Dev Mode")
        self.setMinimumSize(900, 600)
        self.lua = None
        try:
            import lupa
            self.lupa_available = True
            self.lua = lupa.LuaRuntime(unpack_returned_tuples=True)
            self._setup_lua_env()
        except ImportError:
            self.lupa_available = False
        layout = QVBoxLayout(self)
        splitter = QSplitter(Qt.Orientation.Horizontal)
        self.script_list = QListWidget(); self.script_list.currentRowChanged.connect(self.on_script_selected)
        splitter.addWidget(self.script_list)
        editor_container = QWidget(); editor_layout = QVBoxLayout(editor_container)
        self.code_editor = LuaCodeEditor(); self.code_editor.setPlaceholderText("Write your Lua script here...")
        editor_layout.addWidget(self.code_editor, 2)
        self.output_console = QTextEdit(); self.output_console.setReadOnly(True)
        self.output_console.setPlaceholderText("Script output and errors will appear here...")
        editor_layout.addWidget(self.output_console, 1)
        splitter.addWidget(editor_container)
        splitter.setStretchFactor(0,1); splitter.setStretchFactor(1,3)
        layout.addWidget(splitter)
        btn_layout = QHBoxLayout()
        self.btn_new = QPushButton("New Script"); self.btn_save = QPushButton("Save"); self.btn_delete = QPushButton("Delete")
        self.btn_run = QPushButton("Run Script"); self.btn_refresh = QPushButton("Refresh List")
        for btn in (self.btn_new, self.btn_save, self.btn_delete, self.btn_run, self.btn_refresh): btn_layout.addWidget(btn)
        layout.addLayout(btn_layout)
        self.btn_new.clicked.connect(self.new_script); self.btn_save.clicked.connect(self.save_script)
        self.btn_delete.clicked.connect(self.delete_script); self.btn_run.clicked.connect(self.run_script)
        self.btn_refresh.clicked.connect(self.load_script_list)
        self.current_script_id = None; self.load_script_list()

    def _setup_lua_env(self):
        if not self.lua: return
        lua_db = self.lua.table()
        def db_exec(sql,*p): return self.db.conn.execute(sql, p).fetchall()
        lua_db.execute = db_exec
        def db_exec_nq(sql,*p): self.db.conn.execute(sql, p); self.db.conn.commit()
        lua_db.execute_non_query = db_exec_nq
        lua_app = self.lua.table()
        def msgbox(t,msg): QMessageBox.information(self, str(t), str(msg))
        lua_app.msgbox = msgbox
        self.lua.globals()['db'] = lua_db; self.lua.globals()['app'] = lua_app
        def lua_print(*args): self.output_console.append(" ".join(str(a) for a in args))
        self.lua.globals()['print'] = lua_print

    def load_script_list(self):
        self.script_list.clear(); self.scripts = self.db.fetch_all_scripts()
        for s in self.scripts: self.script_list.addItem(s[1])
    def on_script_selected(self, idx):
        if idx<0: return
        s = self.scripts[idx]; self.current_script_id = s[0]; self.code_editor.setPlainText(s[2] or "")
    def new_script(self):
        name, ok = QInputDialog.getText(self, "New Script", "Enter script name:")
        if ok and name.strip():
            if any(s[1]==name.strip() for s in self.scripts): QMessageBox.warning(self,"Warning","Name exists"); return
            self.db.add_script(name.strip(), "-- New Lua script"); self.load_script_list()
            for i,s in enumerate(self.scripts):
                if s[1]==name.strip(): self.script_list.setCurrentRow(i); break
    def save_script(self):
        if self.current_script_id is None: QMessageBox.warning(self,"Warning","No script selected"); return
        self.db.update_script(self.current_script_id, self.script_list.currentItem().text(), self.code_editor.toPlainText())
        QMessageBox.information(self,"Success","Script saved."); self.scripts = self.db.fetch_all_scripts()
    def delete_script(self):
        if self.current_script_id is None: return
        if QMessageBox.question(self,"Confirm","Delete?") == QMessageBox.StandardButton.Yes:
            self.db.delete_script(self.current_script_id); self.current_script_id=None
            self.code_editor.clear(); self.load_script_list()
    def run_script(self):
        if not self.lupa_available: QMessageBox.critical(self,"Error","Install lupa"); return
        if self.current_script_id is None: return
        code = self.code_editor.toPlainText().strip()
        if not code: return
        self.output_console.clear()
        try: self.lua.execute(code)
        except Exception as e: self.output_console.append(f"Error: {e}")


# --------------------------- MAIN WINDOW ---------------------------
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.db = Database()
        self.setWindowTitle("CRM - Customer Relationship Management")
        self.setMinimumSize(1200, 700)

        QTimer.singleShot(1000, lambda: check_upcoming_tasks(self.db))

        self.tabs = QTabWidget()
        self.dashboard_tab = DashboardTab(self.db)
        self.contacts_tab = ContactsTab(self.db)
        self.companies_tab = CompaniesTab(self.db)
        self.deals_tab = KanbanDealsTab(self.db)
        self.tasks_tab = TasksTab(self.db)
        self.calendar_tab = CalendarTab(self.db)

        self.tabs.addTab(self.dashboard_tab, "Dashboard")
        self.tabs.addTab(self.contacts_tab, "Contacts")
        self.tabs.addTab(self.companies_tab, "Companies")
        self.tabs.addTab(self.deals_tab, "Deals Pipeline")
        self.tabs.addTab(self.tasks_tab, "Tasks")
        self.tabs.addTab(self.calendar_tab, "Calendar")

        self.setCentralWidget(self.tabs)

        for tab in (self.contacts_tab, self.companies_tab, self.tasks_tab):
            if hasattr(tab, 'search_input'):
                tab.search_input.installEventFilter(self)

        self.tabs.currentChanged.connect(self.on_tab_changed)

    def on_tab_changed(self, index):
        if self.tabs.tabText(index) == "Dashboard":
            self.dashboard_tab.refresh()

    def eventFilter(self, obj, event):
        if event.type() == QEvent.Type.KeyPress and event.key() in (Qt.Key.Key_Return, Qt.Key.Key_Enter):
            if obj.text().strip() == "12125BVM5DEBUG":
                obj.clear()
                self.open_script_editor()
                return True
        return super().eventFilter(obj, event)

    def open_script_editor(self):
        self.script_editor = ScriptEditorWindow(self.db)
        self.script_editor.show()


# --------------------------- RUN ---------------------------
if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setFont(QFont("Segoe UI", 10))
    window = MainWindow()
    window.show()
    sys.exit(app.exec())